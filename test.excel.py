import openpyxl
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
wb_read = openpyxl.load_workbook('test.xlsx')

sheet = wb_read['Sheet']
barcha_idlar_soni = sheet.max_row


for i in range(1, barcha_idlar_soni + 1):
    id_telegram = ws.cell(row=1, column=1).value
    print(id_telegram)

wb.close()
# wb.save('test.xlsx')
