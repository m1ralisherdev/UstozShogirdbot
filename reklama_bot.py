import logging
from aiogram import Bot, Dispatcher, executor, types
from aiogram.types import Message, ContentType
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import ParseMode
import openpyxl
from openpyxl import Workbook

# -----------------------------------------------------BOT
API_TOKEN = '6705011112:AAEca6XiG4KYGRafdvJtkCug7kIyjLvaGHg'

bot = Bot(token=API_TOKEN, parse_mode='HTML')
dp = Dispatcher(bot, storage=MemoryStorage())
logging.basicConfig(level=logging.INFO)

# ----------------------------------------EXCEL

wb = Workbook()
ws = wb.active
wb_read = openpyxl.load_workbook('database.xlsx')
sheet = wb_read.active

barcha_idlar_soni = sheet.max_row

odamlar = []


# ----------------------------------------DISPATCHER
@dp.message_handler(commands='start')
async def starter(message: Message):
    for i in range(1, barcha_idlar_soni + 1):
        id_telegram = sheet.cell(row=i, column=1).value
        if id_telegram:
            odamlar.append(int(id_telegram))

    if message.from_user.id not in odamlar:
        odamlar.append(message.from_user.id)
        await message.answer("Assalomaleykum Bot ga xush kelibsiz!\n siz uchun bonuslarimiz mavjud")

    else:
        await message.answer(
            "Assalomaleykum Siz botdan oldin ro`yxatdan o`tib bo`lgansiz siz uchun bonuslar mavjud emas!")

    for i in range(1, len(odamlar) + 1):
        ws[f'A{i}'] = odamlar[i - 1]
    wb.save('database.xlsx')

if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True)
